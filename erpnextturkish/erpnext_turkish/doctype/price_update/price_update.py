# Copyright (c) 2025, Logedosoft Business Solutions and contributors
# For license information, please see license.txt

# import frappe

import frappe
import base64
from io import BytesIO
from openpyxl import Workbook, load_workbook
from frappe import _
from frappe.model.document import Document

class PriceUpdate(Document):
    def validate(self):
        pass

def read_xlsx_as_dict(file):
    """Excel dosyasını dictionary olarak okur"""
    wb = load_workbook(file, data_only=True)
    sheet = wb.active
    data = []
    headers = []
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(cell).strip() if cell else '' for cell in row]
        else:
            row_data = {}
            for key, cell in zip(headers, row):
                row_data[str(key).strip()] = str(cell).strip() if cell is not None else ''
            data.append(row_data)
    return data

@frappe.whitelist()
def get_price_update_template():
    """Fiyat güncelleme şablonu Excel dosyası oluşturur"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Fiyat Güncelleme Şablonu"

    headers = ["Ürün Kodu", "Satış Fiyatı", "Satış Fiyat Listesi", "İskonto", "Alış Fiyatı", "Alış Fiyat Listesi"]
    ws.append(headers)

    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    encoded_file = base64.b64encode(file_stream.read()).decode("utf-8")

    return {
        "filename": "fiyat_guncelleme_sablonu.xlsx",
        "filedata": f"data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{encoded_file}"
    }

@frappe.whitelist()
def process_attachment(docname, attachment):
    """Attachment'tan Excel dosyasını işler"""
    doc = frappe.get_doc("Price Update", docname)

    if not attachment:
        frappe.throw(_("Herhangi bir dosya eklenmemiş."))

    # Dosyayı bul: file_url üzerinden
    file_record = frappe.get_all("File", filters={"file_url": attachment}, fields=["name"], limit=1)
    if not file_record:
        frappe.throw(_("Dosya bulunamadı: {0}").format(attachment))

    file_doc = frappe.get_doc("File", file_record[0].name)
    file_content = file_doc.get_content()

    # Geri kalanı aynı kalabilir (okuma, işleme vs)
    excel_stream = BytesIO(file_content)
    rows = read_xlsx_as_dict(excel_stream)
    
    results = {
        "success": 0,
        "updated": 0,
        "created": 0,
        "errors": [],
        "failed_items": []
    }

    for row in rows:
        item_code = row.get("Ürün Kodu") or row.get("Item Code") or row.get("item_code")
        selling_price = row.get("Satış Fiyatı") or row.get("Selling Price") or row.get("selling_price")
        selling_price_list = row.get("Satış Fiyat Listesi") or row.get("Selling Price List") or row.get("selling_price_list")
        buying_price = row.get("Alış Fiyatı") or row.get("Buying Price") or row.get("buying_price")
        buying_price_list = row.get("Alış Fiyat Listesi") or row.get("Buying Price List") or row.get("buying_price_list")

        if not item_code:
            results["errors"].append(f"Ürün Kodu boş: {row}")
            continue

        if not frappe.db.exists("Item", item_code):
            results["errors"].append(f"Ürün bulunamadı: {item_code}")
            results["failed_items"].append(item_code)
            continue

        # Satış fiyatı işle
        if selling_price and selling_price_list:
            try:
                selling_price = float(str(selling_price).replace(",", "."))
                
                if not frappe.db.exists("Price List", selling_price_list):
                    results["errors"].append(f"Satış fiyat listesi bulunamadı: {selling_price_list}")
                    results["failed_items"].append(f"{item_code} (Satış Fiyat Listesi: {selling_price_list})")
                    continue

                existing_price = frappe.db.get_value("Item Price", {
                    "item_code": item_code,
                    "price_list": selling_price_list
                }, "name")

                if existing_price:
                    price_doc = frappe.get_doc("Item Price", existing_price)
                    price_doc.price_list_rate = selling_price
                    price_doc.save()
                    results["updated"] += 1
                else:
                    price_doc = frappe.get_doc({
                        "doctype": "Item Price",
                        "item_code": item_code,
                        "price_list": selling_price_list,
                        "price_list_rate": selling_price,
                        "currency": frappe.db.get_value("Price List", selling_price_list, "currency") or "TRY"
                    })
                    price_doc.insert()
                    results["created"] += 1

                results["success"] += 1

            except Exception as e:
                results["errors"].append(f"Satış fiyatı hatası: {item_code} - {str(e)}")
                results["failed_items"].append(item_code)

        # Alış fiyatı işle
        if buying_price and buying_price_list:
            try:
                buying_price = float(str(buying_price).replace(",", "."))
                
                if not frappe.db.exists("Price List", buying_price_list):
                    results["errors"].append(f"Alış fiyat listesi bulunamadı: {buying_price_list}")
                    results["failed_items"].append(f"{item_code} (Alış Fiyat Listesi: {buying_price_list})")
                    continue

                existing_price = frappe.db.get_value("Item Price", {
                    "item_code": item_code,
                    "price_list": buying_price_list
                }, "name")

                if existing_price:
                    price_doc = frappe.get_doc("Item Price", existing_price)
                    price_doc.price_list_rate = buying_price
                    price_doc.save()
                    results["updated"] += 1
                else:
                    price_doc = frappe.get_doc({
                        "doctype": "Item Price",
                        "item_code": item_code,
                        "price_list": buying_price_list,
                        "price_list_rate": buying_price,
                        "currency": frappe.db.get_value("Price List", buying_price_list, "currency") or "TRY"
                    })
                    price_doc.insert()
                    results["created"] += 1

                results["success"] += 1

            except Exception as e:
                results["errors"].append(f"Alış fiyatı hatası: {item_code} - {str(e)}")
                results["failed_items"].append(item_code)

    doc.total_processed = results["success"] + len(results["errors"])
    doc.successful_updates = results["updated"]
    doc.new_prices_created = results["created"]
    doc.errors = "\n".join([str(e) for e in results["errors"]]) if results["errors"] else ""

    if results["failed_items"]:
        failed_items_text = "Oluşturulamayan/Güncellenemeyen Ürünler:\n" + "\n".join([str(item) for item in results["failed_items"]])
        if doc.errors:
            doc.errors += "\n\n" + failed_items_text
        else:
            doc.errors = failed_items_text

    if results["success"] > 0 and len(results["errors"]) == 0:
        doc.status = "Tamamlandı"
    elif results["success"] > 0 and len(results["errors"]) > 0:
        doc.status = "Isleniyor"
    elif results["success"] == 0 and len(results["errors"]) > 0:
        doc.status = "Başarısız"
    else:
        doc.status = "Draft"

    doc.save()

    if results["success"] > 0 and len(results["errors"]) == 0:
        frappe.msgprint(
            msg=f"Fiyat güncelleme tamamlandı.<br>Güncellenen: {results['updated']}<br>Oluşturulan: {results['created']}<br>Toplam Başarılı: {results['success']}",
            alert=True,
            indicator="green"
        )

    return results